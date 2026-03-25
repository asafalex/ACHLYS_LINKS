"""
ACHLYS Report Uploader
Reads the daily Excel file from OneDrive and uploads to Supabase.
Run daily via Windows Task Scheduler or manually.
"""

import base64
import os
import sys
import tempfile
from datetime import date
from openpyxl import load_workbook
from supabase import create_client

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ===== CONFIG =====
SUPABASE_URL = "https://exrmnluywpzbhqjngjfz.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImV4cm1ubHV5d3B6Ymhxam5namZ6Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzQyNzY2MjMsImV4cCI6MjA4OTg1MjYyM30.DNqpguiybLmt0Oz8PMGRcx_U6dMpAMt7fkdKSDal5d8"

# OneDrive path - adjust if needed
ONEDRIVE_PATH = os.path.expanduser(r"~\OneDrive - Mobile Brain\report_ai_mobile\FINAL")

# SharePoint shared folder link (fallback when local file is locked)
SHAREPOINT_FOLDER_URL = "https://mobilebrain-my.sharepoint.com/:f:/g/personal/asaf_mobile-brain_net/IgDrqGhW4YQUQZt2XzufgeKbAcjdlOg7UTRGU6045f6luas?e=2egc1d"

# Azure AD credentials for Graph API (set these as environment variables or fill directly)
# To register: portal.azure.com → App registrations → New registration
# Permissions needed: Files.Read.All (Application)
AZURE_TENANT_ID = os.environ.get("AZURE_TENANT_ID", "")
AZURE_CLIENT_ID = os.environ.get("AZURE_CLIENT_ID", "")
AZURE_CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")


def get_graph_token():
    """Get Microsoft Graph API token using client credentials flow."""
    if not (AZURE_TENANT_ID and AZURE_CLIENT_ID and AZURE_CLIENT_SECRET):
        return None
    token_url = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(token_url, data={
        "grant_type": "client_credentials",
        "client_id": AZURE_CLIENT_ID,
        "client_secret": AZURE_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }, timeout=30)
    if resp.status_code == 200:
        return resp.json().get("access_token")
    print(f"WARNING: Failed to get Graph token: {resp.status_code} {resp.text[:200]}")
    return None


def download_from_sharepoint(filename):
    """Download a file from the shared SharePoint folder via Microsoft Graph API."""
    if not HAS_REQUESTS:
        print("WARNING: 'requests' library not installed, cannot use SharePoint fallback")
        return None

    # Encode the sharing URL for Graph API
    encoded = base64.b64encode(SHAREPOINT_FOLDER_URL.encode()).decode()
    encoded = encoded.rstrip("=").replace("/", "_").replace("+", "-")
    shares_token = f"u!{encoded}"

    api_url = f"https://graph.microsoft.com/v1.0/shares/{shares_token}/root/children"
    print(f"Fetching SharePoint folder listing...")

    headers = {}
    token = get_graph_token()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    else:
        print("WARNING: No Azure credentials configured, trying anonymous access...")

    resp = requests.get(api_url, headers=headers, timeout=30)

    if resp.status_code != 200:
        print(f"WARNING: SharePoint API returned {resp.status_code}: {resp.text[:200]}")
        return None

    items = resp.json().get("value", [])
    for item in items:
        if item.get("name") == filename:
            download_url = item.get("@microsoft.graph.downloadUrl")
            if not download_url:
                print("WARNING: No downloadUrl found for file")
                return None
            print(f"Downloading {filename} from SharePoint...")
            file_resp = requests.get(download_url, timeout=60)
            if file_resp.status_code != 200:
                print(f"WARNING: Download failed with status {file_resp.status_code}")
                return None
            tmp_path = os.path.join(tempfile.gettempdir(), filename)
            with open(tmp_path, "wb") as f:
                f.write(file_resp.content)
            print(f"Downloaded to temp: {tmp_path}")
            return tmp_path

    print(f"WARNING: '{filename}' not found in SharePoint folder")
    return None

# Column mapping (Excel header -> DB column)
COL_MAP = {
    "Month": "month",
    "Platform": "platform",
    "Source": "source",
    "Campaign Name": "campaign_name",
    "GEO": "geo",
    "Manager": "manager",
    "Brand": "brand",
    "Spend": "spend",
    "Agency Fee": "agency_fee",
    "Real Spend": "real_spend",
    "Installs": "installs",
    "Clicks": "clicks",
    "Reg": "reg",
    "FTD": "ftd",
    "eCPA": "ecpa",
    "Income": "income",
    "Profit": "profit",
    "ROI": "roi",
    "Total Commission (ZAR)": "total_commission_zar",
}


def find_todays_file():
    """Find today's Excel file in OneDrive."""
    today_str = date.today().isoformat()  # 2026-03-25
    filename = f"ACHLYS_Summary_{today_str}.xlsx"
    filepath = os.path.join(ONEDRIVE_PATH, filename)

    if os.path.exists(filepath):
        return filepath

    # Also check Downloads/AIAI folder
    alt_path = os.path.join(os.path.expanduser("~"), "Downloads", "AIAI", filename)
    if os.path.exists(alt_path):
        return alt_path

    # List available files
    print(f"File not found: {filename}")
    print(f"Checked: {ONEDRIVE_PATH}")
    if os.path.exists(ONEDRIVE_PATH):
        files = [f for f in os.listdir(ONEDRIVE_PATH) if f.startswith("ACHLYS_Summary")]
        if files:
            print(f"Available files: {', '.join(sorted(files)[-5:])}")
    return None


def parse_excel(filepath):
    """Parse the Excel file and return rows as dicts."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Find header row
    header_row = None
    subtitle = ""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        values = [cell.value for cell in row]
        if any(v and str(v).strip() in ("Month", "Platform", "Brand") for v in values):
            header_row = row
            break
        # Capture subtitle
        if values[0] and "Exchange" in str(values[0]):
            subtitle = str(values[0])

    if not header_row:
        print("ERROR: Could not find header row in Excel file")
        return [], ""

    # Map headers to column indices
    headers = {}
    for cell in header_row:
        val = str(cell.value).strip() if cell.value else ""
        if val in COL_MAP:
            headers[cell.column - 1] = COL_MAP[val]

    # Parse data rows
    rows = []
    started = False
    for row in ws.iter_rows(values_only=True):
        if not started:
            if any(v and str(v).strip() in ("Month", "Platform") for v in row):
                started = True
            continue

        # Skip empty rows and TOTAL row
        if not row or not any(row):
            continue
        if row[0] and str(row[0]).strip().upper() == 'TOTAL':
            continue

        record = {"report_date": date.today().isoformat(), "subtitle": subtitle}
        for col_idx, db_col in headers.items():
            val = row[col_idx] if col_idx < len(row) else None

            if val is None:
                val = 0 if db_col in ("spend", "agency_fee", "real_spend", "installs",
                                       "clicks", "reg", "ftd", "ecpa", "income",
                                       "profit", "roi", "total_commission_zar") else ""
            elif db_col in ("installs", "clicks", "reg", "ftd"):
                try:
                    val = int(float(val))
                except (ValueError, TypeError):
                    val = 0
            elif db_col in ("spend", "agency_fee", "real_spend", "ecpa",
                           "income", "profit", "roi", "total_commission_zar"):
                try:
                    val = round(float(val), 6)
                except (ValueError, TypeError):
                    val = 0
            else:
                val = str(val).strip() if val else ""

            record[db_col] = val

        rows.append(record)

    wb.close()
    return rows, subtitle


def upload_to_supabase(rows):
    """Upload rows to Supabase, replacing today's data."""
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    today_str = date.today().isoformat()

    # Delete existing data for today
    print(f"Deleting old data for {today_str}...")
    sb.table("report_data").delete().eq("report_date", today_str).execute()

    # Insert new data in batches of 50
    batch_size = 50
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        result = sb.table("report_data").insert(batch).execute()
        print(f"  Uploaded rows {i + 1} to {min(i + batch_size, len(rows))}")

    print(f"SUCCESS: {len(rows)} rows uploaded for {today_str}")


def main():
    print("=" * 50)
    print("ACHLYS Report Uploader")
    print(f"Date: {date.today().isoformat()}")
    print("=" * 50)

    # Allow custom file path as argument
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        if not os.path.exists(filepath):
            print(f"ERROR: File not found: {filepath}")
            return
        print(f"File: {filepath}")
        print("Parsing Excel...")
        rows, subtitle = parse_excel(filepath)
    else:
        today_str = date.today().isoformat()
        filename = f"ACHLYS_Summary_{today_str}.xlsx"

        # 1st priority: SharePoint via Graph API
        print("Trying SharePoint (primary)...")
        filepath = download_from_sharepoint(filename)

        if not filepath:
            # 2nd priority: robocopy from OneDrive to temp
            local_path = find_todays_file()
            if local_path:
                print(f"WARNING: SharePoint unavailable. Trying robocopy from OneDrive...")
                tmp_dir = tempfile.gettempdir()
                tmp_path = os.path.join(tmp_dir, filename)
                result = __import__('subprocess').run(
                    ['robocopy', os.path.dirname(local_path), tmp_dir, filename, '/R:0', '/W:0'],
                    capture_output=True, timeout=30
                )
                if os.path.exists(tmp_path):
                    print(f"Copied to temp: {tmp_path}")
                    filepath = tmp_path
                else:
                    print(f"WARNING: robocopy failed (exit {result.returncode}). Falling back to direct OneDrive read...")
                    filepath = local_path
            else:
                print("ERROR: SharePoint unavailable and no local OneDrive file found")
                return

        if not filepath:
            print("ERROR: No file found for today")
            return

        print(f"File: {filepath}")
        print("Parsing Excel...")
        rows, subtitle = parse_excel(filepath)

    if not rows:
        print("ERROR: No data rows found")
        return

    print(f"Found {len(rows)} data rows")
    if subtitle:
        print(f"Subtitle: {subtitle[:80]}...")

    upload_to_supabase(rows)
    print("Done!")


if __name__ == "__main__":
    main()
